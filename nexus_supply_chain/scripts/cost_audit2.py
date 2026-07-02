import frappe
import openpyxl
import openpyxl.styles
import openpyxl.utils

# ---- CONFIG: adjust dates as needed ----
FROM_DATE = "2026-06-01"
TO_DATE = "2026-06-30"
COMPANY = None  # set to e.g. "Crystal Custom Ltd" to filter, or leave None for all companies

def get_system_cogs(from_date, to_date, company=None):
    print(f"⏳ Pulling recorded COGS from GL Entry + Stock Ledger Entry between {from_date} and {to_date}...")

    company_filter = ""
    params = [from_date, to_date]
    if company:
        company_filter = "AND gle.company = %s"
        params.append(company)

    rows = frappe.db.sql(f"""
        SELECT
            gle.posting_date,
            gle.account,
            gle.voucher_type,
            gle.voucher_no,
            gle.company,
            gle.debit AS cogs_amount,
            sle.item_code,
            it.item_name,
            it.item_group,
            sle.actual_qty,
            sle.valuation_rate,
            sle.stock_value_difference,
            sle.warehouse
        FROM `tabGL Entry` gle
        LEFT JOIN `tabStock Ledger Entry` sle
            ON sle.voucher_no = gle.voucher_no
            AND sle.voucher_type = gle.voucher_type
            AND sle.company = gle.company
        LEFT JOIN `tabItem` it
            ON it.name = sle.item_code
        JOIN `tabAccount` acc
            ON acc.name = gle.account
        WHERE acc.account_type = 'Cost of Goods Sold'
          AND gle.is_cancelled = 0
          AND gle.posting_date BETWEEN %s AND %s
          {company_filter}
          AND sle.item_code IS NOT NULL
        ORDER BY gle.posting_date DESC, gle.voucher_no
    """, tuple(params), as_dict=True)

    print(f"✅ Pulled {len(rows)} line items.")
    return rows


def aggregate_by_item(rows):
    """Roll up per-item totals: total qty delivered, total recorded COGS, weighted avg valuation rate."""
    agg = {}
    for r in rows:
        key = r.item_code
        if key not in agg:
            agg[key] = {
                "item_code": r.item_code,
                "item_name": r.item_name,
                "item_group": r.item_group,
                "total_qty": 0.0,
                "total_cogs": 0.0,
            }
        agg[key]["total_qty"] += frappe.utils.flt(r.qty) * -1 if frappe.utils.flt(r.qty) < 0 else frappe.utils.flt(r.qty)
        agg[key]["total_cogs"] += frappe.utils.flt(r.cogs_amount)

    for key, v in agg.items():
        v["weighted_avg_rate"] = (v["total_cogs"] / v["total_qty"]) if v["total_qty"] else 0.0

    return agg


def export_system_cogs_report(from_date, to_date, company=None):
    rows = get_system_cogs(from_date, to_date, company)
    if not rows:
        print("⚠️ No COGS GL entries found for this range. Check your date range / account_type setup.")
        return

    agg = aggregate_by_item(rows)

    wb = openpyxl.Workbook()

    # --- Sheet 1: Per-transaction detail ---
    ws_detail = wb.active
    ws_detail.title = "COGS Detail"

    font_header = openpyxl.styles.Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    fill_navy = openpyxl.styles.PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    detail_headers = ["Posting Date", "Voucher Type", "Voucher No", "Item Code", "Item Name",
                       "Item Group", "Warehouse", "Qty", "Valuation Rate", "COGS Amount", "Company"]
    for idx, text in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=1, column=idx, value=text)
        cell.font, cell.fill = font_header, fill_navy

    for i, r in enumerate(rows, start=2):
        ws_detail.cell(row=i, column=1, value=str(r.posting_date))
        ws_detail.cell(row=i, column=2, value=r.voucher_type)
        ws_detail.cell(row=i, column=3, value=r.voucher_no)
        ws_detail.cell(row=i, column=4, value=r.item_code)
        ws_detail.cell(row=i, column=5, value=r.item_name)
        ws_detail.cell(row=i, column=6, value=r.item_group)
        ws_detail.cell(row=i, column=7, value=r.warehouse)
        ws_detail.cell(row=i, column=8, value=frappe.utils.flt(r.actual_qty))
        c_val = ws_detail.cell(row=i, column=9, value=frappe.utils.flt(r.valuation_rate))
        c_val.number_format = "Sh #,##0.00"
        c_cogs = ws_detail.cell(row=i, column=10, value=frappe.utils.flt(r.cogs_amount))
        c_cogs.number_format = "Sh #,##0.00"
        ws_detail.cell(row=i, column=11, value=r.company)

    # --- Sheet 2: Per-item rollup ---
    ws_agg = wb.create_sheet(title="COGS Per Item Rollup")
    agg_headers = ["Item Code", "Item Name", "Item Group", "Total Qty Delivered",
                   "Total Recorded COGS", "Weighted Avg Valuation Rate"]
    for idx, text in enumerate(agg_headers, 1):
        cell = ws_agg.cell(row=1, column=idx, value=text)
        cell.font, cell.fill = font_header, fill_navy

    sorted_agg = sorted(agg.values(), key=lambda v: v["total_cogs"], reverse=True)
    for i, v in enumerate(sorted_agg, start=2):
        ws_agg.cell(row=i, column=1, value=v["item_code"])
        ws_agg.cell(row=i, column=2, value=v["item_name"])
        ws_agg.cell(row=i, column=3, value=v["item_group"])
        ws_agg.cell(row=i, column=4, value=v["total_qty"])
        c_cogs = ws_agg.cell(row=i, column=5, value=v["total_cogs"])
        c_cogs.number_format = "Sh #,##0.00"
        c_rate = ws_agg.cell(row=i, column=6, value=v["weighted_avg_rate"])
        c_rate.number_format = "Sh #,##0.00"

    # Column widths
    for ws, cols in [
        (ws_detail, [13, 16, 22, 16, 30, 18, 18, 12, 16, 16, 20]),
        (ws_agg, [16, 30, 18, 18, 20, 24]),
    ]:
        for idx, width in enumerate(cols, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    export_filename = "/home/cloud/frappe-bench/System_COGS_Report.xlsx"
    wb.save(export_filename)
    print(f"\n🚀 SUCCESS! {len(rows)} detail lines, {len(agg)} unique items.")
    print(f"👉 Saved to: {export_filename}")
    return agg


# Execute
system_cogs_by_item = export_system_cogs_report(FROM_DATE, TO_DATE, COMPANY)
