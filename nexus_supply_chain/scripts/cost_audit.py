# -*- coding: utf-8 -*-
import frappe
import openpyxl
import openpyxl.styles
import openpyxl.utils


class PortfolioCostAuditor:
    def __init__(self):
        print("⏳ Initializing In-Memory Engine and loading system data...")

        prices = frappe.db.sql("""
            SELECT item_code, price_list_rate
            FROM `tabItem Price`
            WHERE price_list = 'Standard Buying' AND buying = 1
        """, as_dict=True)
        self.price_map = {p.item_code: frappe.utils.flt(p.price_list_rate) for p in prices}

        vals = frappe.db.sql("SELECT name, valuation_rate, item_name, item_group FROM `tabItem`", as_dict=True)
        self.item_val_map = {v.name: frappe.utils.flt(v.valuation_rate) for v in vals}
        self.item_meta_map = {v.name: {"item_name": v.item_name, "item_group": v.item_group} for v in vals}

        boms = frappe.db.sql("""
            SELECT name, item, quantity
            FROM `tabBOM`
            WHERE is_active = 1 AND is_default = 1 AND docstatus = 1
        """, as_dict=True)
        self.bom_map = {
            b.item: {"name": b.name, "qty": frappe.utils.flt(b.quantity) or 1.0}
            for b in boms
        }

        self.bom_items_map = {}
        if boms:
            bom_names = tuple(b.name for b in boms)
            bom_items = frappe.db.sql("""
                SELECT parent, item_code, stock_qty, item_name
                FROM `tabBOM Item`
                WHERE parent IN %s
            """, (bom_names,), as_dict=True)
            for bi in bom_items:
                self.bom_items_map.setdefault(bi.parent, []).append(bi)

        self.font_body = openpyxl.styles.Font(name="Segoe UI", size=10, color="000000")
        self.thin_gray = openpyxl.styles.Side(border_style="thin", color="D9D9D9")
        self.border_grid = openpyxl.styles.Border(
            left=self.thin_gray, right=self.thin_gray, top=self.thin_gray, bottom=self.thin_gray
        )

        self.zero_cost_leaves = {}

    def explode_and_trace(self, item_code, ws, row_tracker, depth=0, visited=None, root_item=None):
        if visited is None:
            visited = set()
        if root_item is None:
            root_item = item_code

        indent = "   " * depth
        bullet = "↳ " if depth > 0 else ""

        if item_code in visited:
            return 0.0

        bom_info = self.bom_map.get(item_code)

        if not bom_info:
            cost = self.price_map.get(item_code)
            if not cost:
                cost = self.item_val_map.get(item_code, 0.0)

            if not cost or frappe.utils.flt(cost) == 0.0:
                entry = self.zero_cost_leaves.setdefault(item_code, {"count": 0, "parents": set()})
                entry["count"] += 1
                entry["parents"].add(root_item)

            return cost

        visited.add(item_code)
        bom_name = bom_info["name"]
        bom_yield = bom_info["qty"]
        children = self.bom_items_map.get(bom_name, [])

        total_bom_cost = 0.0

        for child in children:
            child_item = child.item_code
            child_qty = frappe.utils.flt(child.stock_qty)

            child_unit_cost = self.explode_and_trace(
                child_item, ws, row_tracker, depth + 1, visited.copy(), root_item
            )
            extended_cost = child_qty * child_unit_cost
            total_bom_cost += extended_cost

            if ws and row_tracker:
                r = row_tracker["current_row"]

                c_code = ws.cell(row=r, column=3, value=f"{indent}{bullet}{child_item}")
                c_code.font = self.font_body
                c_code.alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")

                c_name = ws.cell(row=r, column=4, value=child.item_name)
                c_name.font = self.font_body

                c_qty = ws.cell(row=r, column=5, value=child_qty)
                c_qty.font = self.font_body
                c_qty.number_format = "#,##0.0000"

                c_cost = ws.cell(row=r, column=6, value=child_unit_cost)
                c_cost.font = self.font_body
                c_cost.number_format = "Sh #,##0.00"
                if not child_unit_cost or frappe.utils.flt(child_unit_cost) == 0.0:
                    c_cost.font = openpyxl.styles.Font(name="Segoe UI", size=10, color="C00000", bold=True)

                c_ext = ws.cell(row=r, column=7, value=extended_cost)
                c_ext.font = self.font_body
                c_ext.number_format = "Sh #,##0.00"

                row_fill = (
                    openpyxl.styles.PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
                    if r % 2 == 0
                    else openpyxl.styles.PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                )
                for col in range(3, 8):
                    ws.cell(row=r, column=col).fill = row_fill
                    ws.cell(row=r, column=col).border = self.border_grid

                ws.row_dimensions[r].height = 18
                row_tracker["current_row"] += 1

        return total_bom_cost / bom_yield

    def run_report(self):
        print("📊 Formatting Workspace...")
        wb = openpyxl.Workbook()

        ws_summary = wb.active
        ws_summary.title = "Portfolio Summary"
        ws_guide = wb.create_sheet(title="BOM Calculation Guide")
        ws_zero = wb.create_sheet(title="Zero-Cost Worklist")

        font_header = openpyxl.styles.Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        fill_navy = openpyxl.styles.PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_red = openpyxl.styles.PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        font_bold = openpyxl.styles.Font(bold=True)

        summary_headers = ["Item Code", "Item Name", "Item Group", "Default BOM Link", "System Valuation", "Theoretical Cost", "Operational Variance"]
        for idx, text in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=4, column=idx, value=text)
            cell.font, cell.fill = font_header, fill_navy

        guide_headers = ["Item Code & Name", "Default BOM Reference", "Component Item Code", "Component Name", "Required Qty", "Unit Cost", "Extended Cost"]
        for idx, text in enumerate(guide_headers, 1):
            cell = ws_guide.cell(row=4, column=idx, value=text)
            cell.font, cell.fill = font_header, fill_navy

        ws_zero['A1'] = "Zero Resolved Cost — Raw Material / BIP Worklist"
        ws_zero['A1'].font = openpyxl.styles.Font(name="Segoe UI", size=16, bold=True, color="1F4E78")
        ws_zero['A2'] = "Leaf-node items with no Standard Buying price AND no Item valuation_rate"
        ws_zero['A2'].font = openpyxl.styles.Font(name="Segoe UI", size=10, italic=True, color="595959")

        zero_headers = ["Item Code", "Item Name", "Item Group", "Times Hit Across BOMs", "Affected Finished Goods (root items)"]
        for idx, text in enumerate(zero_headers, 1):
            cell = ws_zero.cell(row=4, column=idx, value=text)
            cell.font, cell.fill = font_header, fill_red

        fg_bounds = frappe.db.get_value("Item Group", "Finished Goods", ["lft", "rgt"], as_dict=True)
        if not fg_bounds:
            print("❌ Error: 'Finished Goods' Item Group not found.")
            return

        items = frappe.db.sql("""
            SELECT i.name as item_code, i.item_name, i.item_group, i.valuation_rate as static_item_val
            FROM `tabItem` i
            JOIN `tabItem Group` ig ON i.item_group = ig.name
            WHERE i.disabled = 0
              AND ig.lft >= %s AND ig.rgt <= %s
        """, (fg_bounds.lft, fg_bounds.rgt), as_dict=True)

        summary_row = 5
        guide_tracker = {"current_row": 5}
        items_with_boms = 0

        for it in items:
            bom_info = self.bom_map.get(it.item_code)
            if not bom_info:
                continue

            items_with_boms += 1
            bom_no = bom_info["name"]
            system_val = frappe.utils.flt(it.static_item_val)

            g_row = guide_tracker["current_row"]
            ws_guide.cell(row=g_row, column=1, value=f"{it.item_code} - {it.item_name}").font = font_bold
            ws_guide.cell(row=g_row, column=2, value=bom_no).font = font_bold
            guide_tracker["current_row"] += 2

            print(f"📦 Exploding: {it.item_code}")
            theoretical_cost = self.explode_and_trace(
                it.item_code, ws_guide, guide_tracker, depth=0, root_item=it.item_code
            )
            variance_val = theoretical_cost - system_val

            g_row = guide_tracker["current_row"]
            ws_guide.cell(row=g_row, column=3, value="Theoretical Rollup Cost:").font = font_bold
            ws_guide.cell(row=g_row, column=6, value=theoretical_cost).number_format = "Sh #,##0.00"
            guide_tracker["current_row"] += 2

            ws_summary.cell(row=summary_row, column=1, value=it.item_code)
            ws_summary.cell(row=summary_row, column=2, value=it.item_name)
            ws_summary.cell(row=summary_row, column=3, value=it.item_group)
            ws_summary.cell(row=summary_row, column=4, value=bom_no)

            c_sys = ws_summary.cell(row=summary_row, column=5, value=system_val)
            c_sys.number_format = "Sh #,##0.00"

            c_theo = ws_summary.cell(row=summary_row, column=6, value=theoretical_cost)
            c_theo.number_format = "Sh #,##0.00"

            c_var = ws_summary.cell(row=summary_row, column=7, value=variance_val)
            c_var.number_format = "Sh #,##0.00"
            c_var.font = openpyxl.styles.Font(color="C00000" if variance_val > 0 else "385723", bold=True)

            summary_row += 1

        zero_row = 5
        sorted_zero_items = sorted(
            self.zero_cost_leaves.items(), key=lambda kv: kv[1]["count"], reverse=True
        )
        for item_code, info in sorted_zero_items:
            meta = self.item_meta_map.get(item_code, {})
            ws_zero.cell(row=zero_row, column=1, value=item_code)
            ws_zero.cell(row=zero_row, column=2, value=meta.get("item_name", ""))
            ws_zero.cell(row=zero_row, column=3, value=meta.get("item_group", ""))
            ws_zero.cell(row=zero_row, column=4, value=info["count"])
            ws_zero.cell(row=zero_row, column=5, value=", ".join(sorted(info["parents"])))

            row_fill = (
                openpyxl.styles.PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
                if zero_row % 2 == 0
                else openpyxl.styles.PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            )
            for col in range(1, 6):
                ws_zero.cell(row=zero_row, column=col).fill = row_fill
                ws_zero.cell(row=zero_row, column=col).font = self.font_body
                ws_zero.cell(row=zero_row, column=col).border = self.border_grid
            zero_row += 1

        print(f"\n🔎 Zero-cost leaf items found: {len(self.zero_cost_leaves)}")

        for ws, cols in [
            (ws_summary, [20, 35, 20, 25, 18, 18, 20]),
            (ws_guide, [35, 25, 22, 35, 15, 16, 16]),
            (ws_zero, [20, 35, 20, 18, 60]),
        ]:
            for idx, width in enumerate(cols, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

        export_filename = "/home/cloud/frappe-bench/True_Market_Cost_Audit.xlsx"
        wb.save(export_filename)
        print(f"\n🚀 SUCCESS! Compiled {items_with_boms} Finished Goods.")
        print(f"👉 Saved to: {export_filename}")


def run():
    auditor = PortfolioCostAuditor()
    auditor.run_report()
